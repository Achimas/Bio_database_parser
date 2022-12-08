import requests
from openpyxl import Workbook
from bs4 import BeautifulSoup as bs

def open_excel(excel_name):
	wb = Workbook(excel_name)
	ws = wb.active
	all_names = []
	for row in ws.iter_rows(values_only=True):  
		all_names.append(row['B'])  # to do
	wb.save(f'_{excel_name}')
	return all_names

def search_objects(all_names):
	site_url = "..."  # to do
	r = request.get(site_url)  # to do
	with open(r) as r:
		soup = BeautifulSoup(r, 'html.parser')  # to do
	parents_to_xls = []
	for name in all_names:
		link = soup.find(name)  # to do
		all_parents = get_parents(name, link)
		parents_to_xls.append(all_parents)
	return parents_to_xls

def get_parents(name, link):
	all_parents_list = []
	parent_name_now = {}
	parent_name_now[name] = link
	while True:
		adding_parent = {}
		for key in parent_name_now.keys():
			r = request.get(parent_name_now[key])  # to do
			soup = BeautifulSoup(r, 'html.parser')  # to do
			link_name = soup.find("a", class_='...').text #class = '' как-то там называющийся класс-родитель
			link = soup.find("a", class_='...') # .get (href) # class = '' как-то там называющийся класс-родитель
			adding_parent[link_name] = link
		if 'pathways' in adding_parent.keys():
			break
		all_parents_list.append(', '.join(adding_parent.keys()) if len(adding_parent.keys())>1 else adding_parent.keys()[0])
		parent_name_now = adding_parent
	return all_parents_list
			
def adding_parents_excel (parents_to_xls, excel_name):
	wb = Workbook(excel_name)
	ws = wb.active
	ws.insert_cols(1) # вставляем 5 колонок перед основным названием
	ws.insert_cols(1)
	ws.insert_cols(1)
	ws.insert_cols(1)
	ws.insert_cols(1)
	cells_symbol = ['B', 'C', 'D', 'E', 'F']
	for i in range(len(ws.rows)-1):
		for j in range(5):
			cell = f'{cells_symbol}{i+2}'
			ws[cell]=parents_to_xls[i][j]
	wb.save(f'v2_{excel_name}')
